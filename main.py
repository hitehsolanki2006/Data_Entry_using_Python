from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('800x500+300+200')
root.resizable(False,False)
root.configure(bg="#6666ff")

file=pathlib.Path('backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active

    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"
    
    file.save(['backend_data.xlsx'])


def submit():
    
    name=namevalue.get()
    contact=contactEntry.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)
    
    file=openpyxl.load_workbook('backend_data.xlsx')
    sheet=file.active

    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    
    file.save(r'backend_data.xlsx')
    
    messagebox.showinfo('info','detail added!')
    namevalue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)
    

def clear():
    namevalue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)
    
    



#icon
icon_image=PhotoImage(file="hunter_1.png")
root.iconphoto(False,icon_image)

#heading
Label(root,text="Please fill out this Entry from",font="arial 20",bg="#6666ff",fg="#fff").place(x=250,y=20)

#lable
Label(root,text="Name",font=23,bg="#6666ff",fg="#fff").place(x=50,y=100)
Label(root,text="Contact No",font=23,bg="#6666ff",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#6666ff",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#6666ff",fg="#fff").place(x=380,y=200)
Label(root,text="Address",font=23,bg="#6666ff",fg="#fff").place(x=50,y=250)


#Entry
namevalue=StringVar()
contactValue=StringVar()
AgeValue=StringVar()


nameEntry=Entry(root,textvariable=namevalue,width=45,bd=2,font=20)
contactEntry=Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry=Entry(root,textvariable=AgeValue,width=15,font=20)
#after sometime change thw width
#Gender
gender_combobox=Combobox(root,values=['Male','Female'],font='arial 14',state='r',width=14)
gender_combobox.place(x=470,y=200)
gender_combobox.set('Male')

addressEntry=Text(root,width=50,height=4,bd=4)




nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

Button(root,text="Submit",bg="#3333ff",fg="white",width=15,height=2,command=submit).place(x=120,y=390)
Button(root,text="Clear",bg="#3333ff",fg="white",width=15,height=2,command=clear).place(x=340,y=390)
Button(root,text="Exit",bg="#3333ff",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=560,y=390)


root.mainloop()